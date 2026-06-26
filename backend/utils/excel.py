"""
excel.py
--------
Generate an Excel attendance report for a course.
"""
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from models import AttendanceSession, AttendanceRecord


def generate_attendance_excel(course) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"

    # ── Color palette ──
    HEADER_FILL = PatternFill("solid", fgColor="1A1A2E")
    DATE_FILL   = PatternFill("solid", fgColor="16213E")
    GREEN_FILL  = PatternFill("solid", fgColor="0F9B58")
    RED_FILL    = PatternFill("solid", fgColor="E63946")
    WHITE_FONT  = Font(color="FFFFFF", bold=True)
    THIN_BORDER = Border(
        left=Side(style="thin", color="2D2D44"),
        right=Side(style="thin", color="2D2D44"),
        top=Side(style="thin", color="2D2D44"),
        bottom=Side(style="thin", color="2D2D44"),
    )

    sessions = (
        AttendanceSession.query
        .filter_by(course_id=course.id, status="confirmed")
        .order_by(AttendanceSession.date)
        .all()
    )

    # Collect all enrolled students
    enrollments = list(course.enrollments)
    students = [(e.student, e.student.user) for e in enrollments]

    # ── Title Row ──
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3 + len(sessions))
    title_cell = ws.cell(row=1, column=1,
                         value=f"{course.name} ({course.code}) — Attendance Report")
    title_cell.font      = Font(bold=True, size=14, color="FFFFFF")
    title_cell.fill      = HEADER_FILL
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # ── Column headers ──
    headers = ["#", "Student ID", "Name"] + [str(s.date) for s in sessions] + ["Total", "Total Sessions", "%"]
    col_count = len(headers)
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font      = WHITE_FONT
        cell.fill      = DATE_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border    = THIN_BORDER

    # ── Student rows ──
    for row_idx, (student, user) in enumerate(students, start=3):
        ws.cell(row=row_idx, column=1, value=row_idx - 2).alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=2, value=student.student_id).alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=3, value=user.name)

        attended = 0
        for col_idx, sess in enumerate(sessions, start=4):
            rec = AttendanceRecord.query.filter_by(
                session_id=sess.id, student_id=student.id
            ).first()
            present = rec.present if rec else False
            cell = ws.cell(row=row_idx, column=col_idx, value="P" if present else "A")
            cell.fill      = GREEN_FILL if present else RED_FILL
            cell.font      = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.border    = THIN_BORDER
            if present:
                attended += 1

        total_col = 4 + len(sessions)
        pct = round(attended / len(sessions) * 100, 1) if sessions else 0
        ws.cell(row=row_idx, column=total_col,     value=attended).alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=total_col + 1, value=len(sessions)).alignment = Alignment(horizontal="center")
        pct_cell = ws.cell(row=row_idx, column=total_col + 2, value=f"{pct}%")
        pct_cell.alignment = Alignment(horizontal="center")
        pct_cell.font = Font(
            color="0F9B58" if pct >= 75 else "E63946", bold=True
        )

    # ── Column widths ──
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 25
    for i in range(4, col_count + 1):
        ws.column_dimensions[get_column_letter(i)].width = 13

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
